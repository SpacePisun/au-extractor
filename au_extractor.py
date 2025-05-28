import sys
import os
import re
import openpyxl
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QLabel, QLineEdit, QPushButton,
                             QFileDialog, QMessageBox, QCheckBox, QTextEdit,
                             QGroupBox, QRadioButton, QButtonGroup)
from PyQt5.QtCore import Qt, QThread, pyqtSignal


class WorkerThread(QThread):
    """Отдельный поток для выполнения операций извлечения и сохранения данных"""
    finished = pyqtSignal()
    progress = pyqtSignal(str)
    error = pyqtSignal(str)

    def __init__(self, txt_path, excel_path, sheet_name=None, values_count=3):
        super().__init__()
        self.txt_path = txt_path
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        self.values_count = values_count
        self.values = []

    def run(self):
        try:
            # Извлекаем значения из TXT файла
            self.values = self.extract_values_from_txt(self.txt_path)
            self.progress.emit(f"Извлечено {len(self.values)} значений: {self.values}")

            # Записываем в Excel
            self.write_values_to_excel(self.excel_path, self.values, self.sheet_name, self.values_count)
            self.progress.emit(f"Данные успешно записаны в файл {self.excel_path}")

            self.finished.emit()
        except Exception as e:
            self.error.emit(str(e))

    def extract_values_from_txt(self, txt_file_path):
        """
        Извлекает значения перед 'AU' из каждой третьей строки текстового файла.
        """
        values = []
        pattern = r'(\d+(?:\.\d+)?) AU'  # Регулярное выражение для поиска числа перед AU

        with open(txt_file_path, 'r', encoding='utf-8') as file:
            lines = file.readlines()

        # Извлекаем значения из каждой третьей строки (индексы 2, 5, 8, ...)
        for i in range(2, len(lines), 3):  # Начинаем с индекса 2 (третья строка, т.к. индексация с 0)
            if i < len(lines):
                line = lines[i]
                match = re.search(pattern, line)
                if match:
                    values.append(float(match.group(1)))

        return values

    def write_values_to_excel(self, excel_file_path, values, sheet_name=None, values_count=3):
        """
        Записывает значения в Excel файл по указанному шаблону:
        - Для 3 значений: первые три значения в C38:C40, пропуск 4 ячеек, следующие три в C44:C46
        - Для 4 значений: первые четыре значения в C38:C41, пропуск 3 ячеек, следующие четыре в C45:C48
        И так далее.
        """
        # Очищаем путь от лишних символов
        excel_file_path = excel_file_path.strip().strip("'").strip('"')

        # Проверяем и добавляем расширение, если его нет
        if not excel_file_path.lower().endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
            excel_file_path += '.xlsx'

        try:
            # Проверяем, существует ли файл
            if os.path.exists(excel_file_path):
                # Попытка открыть существующий файл
                workbook = openpyxl.load_workbook(excel_file_path)
                self.progress.emit(f"Открыт существующий файл {excel_file_path}")
            else:
                # Если файл не существует, создаем новый
                workbook = openpyxl.Workbook()
                self.progress.emit(f"Создан новый файл {excel_file_path}")
        except Exception as e:
            self.progress.emit(f"Ошибка при открытии/создании файла: {e}")
            self.progress.emit("Создаем новый файл Excel...")
            workbook = openpyxl.Workbook()

        # Определяем рабочий лист
        if sheet_name:
            # Проверяем, существует ли лист с указанным именем
            if sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                self.progress.emit(f"Используется существующий лист '{sheet_name}'")
            else:
                # Создаем новый лист
                worksheet = workbook.create_sheet(sheet_name)
                self.progress.emit(f"Создан новый лист '{sheet_name}'")
        else:
            worksheet = workbook.active
            self.progress.emit(f"Используется активный лист '{worksheet.title}'")

        row = 38  # Начальная строка
        values_index = 0

        # Определяем количество ячеек для пропуска в зависимости от выбранного режима
        if values_count == 3:
            skip_cells = 4  # После 3 значений пропускаем 4 ячейки (как в оригинале)
        else:  # values_count == 4
            skip_cells = 3  # После 4 значений пропускаем 3 ячейки

        self.progress.emit(f"Режим записи: по {values_count} значения с пропуском {skip_cells} ячеек")

        while values_index < len(values):
            # Записываем блок из values_count значений
            for i in range(values_count):
                if values_index < len(values):
                    worksheet[f'C{row}'] = values[values_index]
                    self.progress.emit(f"Записано значение {values[values_index]} в ячейку C{row}")
                    values_index += 1
                    row += 1

            # Пропускаем указанное количество ячеек
            # Учитываем, что row уже увеличен на values_count в предыдущем цикле
            row += skip_cells - 1  # -1 потому что в следующей итерации row снова увеличится

        workbook.save(excel_file_path)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        # Устанавливаем основные параметры окна
        self.setWindowTitle('AU Extractor - Извлечение значений AU')
        self.setGeometry(300, 300, 700, 600)

        # Создаем центральный виджет и основной макет
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        main_layout = QVBoxLayout(self.central_widget)

        # Группа выбора файлов
        file_group = QGroupBox("Выбор файлов")
        files_layout = QVBoxLayout()

        # Строка для выбора TXT файла
        txt_layout = QHBoxLayout()
        txt_layout.addWidget(QLabel("Текстовый файл:"))
        self.txt_path_edit = QLineEdit()
        txt_layout.addWidget(self.txt_path_edit)
        self.txt_browse_btn = QPushButton("Обзор...")
        self.txt_browse_btn.clicked.connect(self.browse_txt_file)
        txt_layout.addWidget(self.txt_browse_btn)
        files_layout.addLayout(txt_layout)

        # Строка для выбора Excel файла
        excel_layout = QHBoxLayout()
        excel_layout.addWidget(QLabel("Excel файл:"))
        self.excel_path_edit = QLineEdit()
        excel_layout.addWidget(self.excel_path_edit)
        self.excel_browse_btn = QPushButton("Обзор...")
        self.excel_browse_btn.clicked.connect(self.browse_excel_file)
        excel_layout.addWidget(self.excel_browse_btn)
        files_layout.addLayout(excel_layout)

        file_group.setLayout(files_layout)
        main_layout.addWidget(file_group)

        # Группа настроек
        settings_group = QGroupBox("Настройки")
        settings_layout = QVBoxLayout()

        # Настройка для выбора листа
        sheet_layout = QHBoxLayout()
        self.use_sheet_checkbox = QCheckBox("Использовать конкретный лист")
        self.use_sheet_checkbox.stateChanged.connect(self.toggle_sheet_name)
        sheet_layout.addWidget(self.use_sheet_checkbox)

        self.sheet_name_edit = QLineEdit()
        self.sheet_name_edit.setEnabled(False)
        sheet_layout.addWidget(self.sheet_name_edit)
        settings_layout.addLayout(sheet_layout)

        # Добавляем настройку для выбора количества значений
        values_count_layout = QVBoxLayout()
        values_count_label = QLabel("Количество значений в блоке:")
        values_count_layout.addWidget(values_count_label)

        # Создаем группу радиокнопок
        self.values_count_group = QButtonGroup()
        radio_layout = QHBoxLayout()

        self.radio_3_values = QRadioButton("3 значения")
        self.radio_3_values.setChecked(True)  # По умолчанию выбрано 3 значения
        self.values_count_group.addButton(self.radio_3_values, 3)
        radio_layout.addWidget(self.radio_3_values)

        self.radio_4_values = QRadioButton("4 значения")
        self.values_count_group.addButton(self.radio_4_values, 4)
        radio_layout.addWidget(self.radio_4_values)

        radio_layout.addStretch()  # Добавляем растяжку для выравнивания влево
        values_count_layout.addLayout(radio_layout)

        # Добавляем пояснительный текст
        explanation_label = QLabel("• 3 значения: пропуск 4 ячеек между блоками\n• 4 значения: пропуск 3 ячеек между блоками")
        explanation_label.setStyleSheet("color: gray; font-size: 10px;")
        values_count_layout.addWidget(explanation_label)

        settings_layout.addLayout(values_count_layout)
        settings_group.setLayout(settings_layout)
        main_layout.addWidget(settings_group)

        # Окно для вывода логов
        log_group = QGroupBox("Журнал выполнения")
        log_layout = QVBoxLayout()
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        log_layout.addWidget(self.log_text)
        log_group.setLayout(log_layout)
        main_layout.addWidget(log_group)

        # Кнопки управления
        buttons_layout = QHBoxLayout()

        self.process_button = QPushButton("Обработать")
        self.process_button.clicked.connect(self.process_files)
        buttons_layout.addWidget(self.process_button)

        self.clear_log_button = QPushButton("Очистить журнал")
        self.clear_log_button.clicked.connect(self.clear_log)
        buttons_layout.addWidget(self.clear_log_button)

        main_layout.addLayout(buttons_layout)

    def browse_txt_file(self):
        """Открывает диалог выбора текстового файла"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Выбрать текстовый файл", "", "Текстовые файлы (*.txt);;Все файлы (*)"
        )
        if file_path:
            self.txt_path_edit.setText(file_path)

    def browse_excel_file(self):
        """Открывает диалог выбора Excel файла"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Выбрать или создать Excel файл", "",
            "Excel файлы (*.xlsx *.xlsm);;Все файлы (*)"
        )
        if file_path:
            self.excel_path_edit.setText(file_path)

    def toggle_sheet_name(self, state):
        """Включает или отключает поле для имени листа"""
        self.sheet_name_edit.setEnabled(state == Qt.Checked)

    def get_selected_values_count(self):
        """Возвращает выбранное количество значений"""
        return self.values_count_group.checkedId()

    def process_files(self):
        """Запускает процесс обработки файлов"""
        # Проверяем, что указаны пути к файлам
        txt_path = self.txt_path_edit.text().strip()
        excel_path = self.excel_path_edit.text().strip()

        if not txt_path:
            QMessageBox.warning(self, "Предупреждение", "Укажите путь к текстовому файлу!")
            return

        if not excel_path:
            QMessageBox.warning(self, "Предупреждение", "Укажите путь к Excel файлу!")
            return

        # Проверяем существование текстового файла
        if not os.path.exists(txt_path):
            QMessageBox.critical(self, "Ошибка", f"Файл {txt_path} не существует!")
            return

        # Определяем имя листа (если нужно)
        sheet_name = None
        if self.use_sheet_checkbox.isChecked():
            sheet_name = self.sheet_name_edit.text().strip()
            if not sheet_name:
                QMessageBox.warning(self, "Предупреждение", "Укажите имя листа!")
                return

        # Получаем выбранное количество значений
        values_count = self.get_selected_values_count()

        # Деактивируем кнопку, чтобы избежать повторных нажатий
        self.process_button.setEnabled(False)
        self.log(f"Начало обработки файлов в режиме {values_count} значений...")

        # Запускаем обработку в отдельном потоке
        self.worker = WorkerThread(txt_path, excel_path, sheet_name, values_count)
        self.worker.progress.connect(self.log)
        self.worker.error.connect(self.show_error)
        self.worker.finished.connect(self.on_processing_finished)
        self.worker.start()

    def on_processing_finished(self):
        """Вызывается по завершении обработки"""
        self.process_button.setEnabled(True)
        self.log("Обработка завершена.")
        QMessageBox.information(self, "Информация", "Обработка файлов успешно завершена!")

    def log(self, message):
        """Добавляет сообщение в журнал"""
        self.log_text.append(message)

    def show_error(self, error_message):
        """Отображает сообщение об ошибке"""
        self.log(f"ОШИБКА: {error_message}")
        QMessageBox.critical(self, "Ошибка", f"Произошла ошибка при обработке: {error_message}")
        self.process_button.setEnabled(True)

    def clear_log(self):
        """Очищает журнал сообщений"""
        self.log_text.clear()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())